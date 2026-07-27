import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def executar_conciliacao():
    # Descobre automaticamente a pasta onde o arquivo 'comparador.py' está salvo
    pasta_do_script = os.path.dirname(os.path.abspath(__file__))
    
    # Monta o caminho completo para os arquivos de dados na mesma pasta
    arquivo_salesrun = os.path.join(pasta_do_script, 'salesrun.csv')
    arquivo_consinco = os.path.join(pasta_do_script, 'consinco.csv')

    # Verifica se os arquivos realmente existem lá
    if not os.path.exists(arquivo_salesrun) or not os.path.exists(arquivo_consinco):
        print(f"Erro: Os arquivos 'salesrun.csv' e 'consinco.csv' não foram encontrados na pasta:\n{pasta_do_script}")
        return

    print("Carregando arquivos...")
    # Lendo os CSVs...
    # (O restante do código continua exatamente igual!)

# --- FUNÇÕES DE LIMPEZA E TRATAMENTO DE DADOS ---

def extrair_id_loja(nome_loja):
    """Extrai apenas o número da loja (ex: 'Loja 10 Ema' -> 10, '12-CRUZEIRO' -> 12)."""
    if pd.isna(nome_loja):
        return None
    match = re.search(r'\d+', str(nome_loja))
    return int(match.group()) if match else None

def limpar_valor_monetario(valor):
    """Trata valores caso venham com pontos, vírgulas e R$ (ex: '157978,84' -> 157978.84)"""
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    
    val_str = str(valor).replace("R$", "").strip()
    # Identifica se há separador de milhar tradicional para limpar corretamente
    if "." in val_str and "," in val_str:
        val_str = val_str.replace(".", "") # Remove pontos de milhar
    val_str = val_str.replace(",", ".")     # Substitui vírgula decimal por ponto
    try:
        return float(val_str)
    except ValueError:
        return 0.0

# --- PROCESSAMENTO PRINCIPAL ---

def executar_conciliacao():
    arquivo_salesrun = 'salesrun.csv'
    arquivo_consinco = 'consinco.csv'

    # Verifica se os arquivos estão na mesma pasta
    if not os.path.exists(arquivo_salesrun) or not os.path.exists(arquivo_consinco):
        print("Erro: Os arquivos 'salesrun.csv' e 'consinco.csv' devem estar na mesma pasta deste script!")
        return

    print("Carregando arquivos...")
    # Lendo os CSVs. O Salesrun costuma vir em codificação ISO/Latin1 devido a acentos.
    df_salesrun = pd.read_csv(arquivo_salesrun, sep=';', names=['Loja', 'Valor'], encoding='latin1')
    df_consinco = pd.read_csv(arquivo_consinco, sep=';', names=['Loja', 'Valor'], encoding='utf-8')

    print("Tratando dados e identificando IDs de lojas...")
    df_salesrun['id_loja'] = df_salesrun['Loja'].apply(extrair_id_loja)
    df_salesrun['venda_salesrun'] = df_salesrun['Valor'].apply(limpar_valor_monetario)

    df_consinco['id_loja'] = df_consinco['Loja'].apply(extrair_id_loja)
    df_consinco['venda_erp'] = df_consinco['Valor'].apply(limpar_valor_monetario)

    print("Fazendo o cruzamento (merge)...")
    df_merged = pd.merge(
        df_salesrun[['id_loja', 'Loja', 'venda_salesrun']],
        df_consinco[['id_loja', 'Loja', 'venda_erp']],
        on='id_loja',
        how='outer',
        suffixes=('_salesrun', '_consinco')
    )

    # Preenche possíveis lojas faltantes com valor 0.0
    df_merged['venda_salesrun'] = df_merged['venda_salesrun'].fillna(0.0)
    df_merged['venda_erp'] = df_merged['venda_erp'].fillna(0.0)
    df_merged['Diferença'] = df_merged['venda_erp'] - df_merged['venda_salesrun']

    # Define o status de batimento
    df_merged['Status'] = df_merged['Diferença'].apply(lambda x: 'OK' if abs(x) < 0.01 else 'DIVERGENTE')
    
    # Ordena para colocar as divergências no topo
    df_merged = df_merged.sort_values(by=['Status', 'id_loja'], ascending=[True, True])

    # --- MONTAGEM DO EXCEL ESTILIZADO ---
    print("Gerando arquivo Excel estilizado...")
    wb = openpyxl.Workbook()

    # Planilha 1: Dashboard / Resumo
    ws_summary = wb.active
    ws_summary.title = "Painel de Controle"
    ws_summary.views.sheetView[0].showGridLines = True

    # Planilha 2: Detalhes
    ws_details = wb.create_sheet(title="Batimento Detalhado")
    ws_details.views.sheetView[0].showGridLines = True

    # Definição de Estilos Visuais
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_divergent = Font(name="Segoe UI", size=10, bold=True, color="C00000")

    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_divergent_row = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_accent_gray = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='A6A6A6'), bottom=Side(style='double', color='1F4E79')
    )

    # --- ABA DETALHADA ---
    headers = ["Cód. Loja", "Nome Loja (Salesrun)", "Venda Salesrun (A)", 
               "Nome Loja (Consinco)", "Venda Consinco (B)", "Diferença (B - A)", "Status"]

    for col_num, header in enumerate(headers, 1):
        cell = ws_details.cell(row=3, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_details.row_dimensions[3].height = 28
    ws_details.cell(row=1, column=1).value = "Batimento de Vendas Diário - Consinco vs. Salesrun"
    ws_details.cell(row=1, column=1).font = font_title
    ws_details.cell(row=2, column=1).value = "Relatório gerado de forma automatizada por script Python"
    ws_details.cell(row=2, column=1).font = font_subtitle

    # Escrevendo os dados na aba detalhada
    row_idx = 4
    for idx, row in df_merged.iterrows():
        r_data = [
            row['id_loja'], row['Loja_salesrun'], row['venda_salesrun'],
            row['Loja_consinco'], row['venda_erp'], row['Diferença'], row['Status']
        ]
        
        is_div = row['Status'] == 'DIVERGENTE'
        row_fill = fill_divergent_row if is_div else (fill_zebra if row_idx % 2 == 0 else None)
        
        for col_num, val in enumerate(r_data, 1):
            cell = ws_details.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
                
            if col_num in [1, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_num == 7:
                    cell.font = font_divergent if is_div else Font(name="Segoe UI", size=10, color="375623", bold=True)
                else:
                    cell.font = font_bold
            elif col_num in [2, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = font_body
            elif col_num in [3, 5, 6]:
                cell.number_format = 'R$ #,##0.00;R$ (#,##0.00);"-"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = font_divergent if (col_num == 6 and is_div) else font_body

        ws_details.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Linha de Totais (Aba Detalhada)
    ws_details.cell(row=row_idx, column=1).value = "TOTAL"
    ws_details.cell(row=row_idx, column=1).font = font_bold
    ws_details.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

    for col_num in range(1, 8):
        ws_details.cell(row=row_idx, column=col_num).border = double_bottom_border
        ws_details.cell(row=row_idx, column=col_num).fill = fill_accent_gray

    ws_details.cell(row=row_idx, column=3).value = f"=SUM(C4:C{row_idx-1})"
    ws_details.cell(row=row_idx, column=3).number_format = 'R$ #,##0.00'
    ws_details.cell(row=row_idx, column=3).font = font_bold

    ws_details.cell(row=row_idx, column=5).value = f"=SUM(E4:E{row_idx-1})"
    ws_details.cell(row=row_idx, column=5).number_format = 'R$ #,##0.00'
    ws_details.cell(row=row_idx, column=5).font = font_bold

    ws_details.cell(row=row_idx, column=6).value = f"=SUM(F4:F{row_idx-1})"
    ws_details.cell(row=row_idx, column=6).number_format = 'R$ #,##0.00'
    ws_details.cell(row=row_idx, column=6).font = font_bold

    # Auto-ajuste de colunas
    for col in ws_details.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3:
                continue
            val_str = str(cell.value or '')
            if 'SUM' in val_str:
                val_str = "R$ 9.999.999,99"
            max_len = max(max_len, len(val_str))
        ws_details.column_dimensions[col_letter].width = max(max_len + 3, 13)

    # --- ABA RESUMO (DASHBOARD) ---
    ws_summary.cell(row=2, column=2).value = "DASHBOARD DE CONCILIAÇÃO"
    ws_summary.cell(row=2, column=2).font = font_title

    # KPI Card 1: Quantidade de Lojas Divergentes
    ws_summary.cell(row=4, column=2).value = "Lojas Divergentes"
    ws_summary.cell(row=4, column=2).font = Font(name="Segoe UI", size=10, color="595959")
    ws_summary.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="center")

    total_div_count = len(df_merged[df_merged['Status'] == 'DIVERGENTE'])
    ws_summary.cell(row=5, column=2).value = total_div_count
    ws_summary.cell(row=5, column=2).font = Font(name="Segoe UI", size=24, bold=True, color="C00000" if total_div_count > 0 else "375623")
    ws_summary.cell(row=5, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws_summary.merge_cells("B4:C4")
    ws_summary.merge_cells("B5:C5")

    # KPI Card 2: Valor Total Divergente
    ws_summary.cell(row=4, column=5).value = "Valor Total Divergente"
    ws_summary.cell(row=4, column=5).font = Font(name="Segoe UI", size=10, color="595959")
    ws_summary.cell(row=4, column=5).alignment = Alignment(horizontal="center", vertical="center")

    ws_summary.cell(row=5, column=5).value = f"=SUM('Batimento Detalhado'!F4:F{row_idx-1})"
    ws_summary.cell(row=5, column=5).number_format = 'R$ #,##0.00'
    ws_summary.cell(row=5, column=5).font = Font(name="Segoe UI", size=16, bold=True, color="C00000" if total_div_count > 0 else "375623")
    ws_summary.cell(row=5, column=5).alignment = Alignment(horizontal="center", vertical="center")

    ws_summary.merge_cells("E4:F4")
    ws_summary.merge_cells("E5:F5")

    # Estilizando os Cards
    card_border_side = Side(style='medium', color='1F4E79')
    for r in range(4, 6):
        for c in range(2, 4):
            cell = ws_summary.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
            top_s = card_border_side if r == 4 else None
            bottom_s = card_border_side if r == 5 else None
            left_s = card_border_side if c == 2 else None
            right_s = card_border_side if c == 3 else None
            cell.border = Border(top=top_s, bottom=bottom_s, left=left_s, right=right_s)

    for r in range(4, 6):
        for c in range(5, 7):
            cell = ws_summary.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
            top_s = card_border_side if r == 4 else None
            bottom_s = card_border_side if r == 5 else None
            left_s = card_border_side if c == 5 else None
            right_s = card_border_side if c == 6 else None
            cell.border = Border(top=top_s, bottom=bottom_s, left=left_s, right=right_s)

    # Tabela Rápida de Erros no Dashboard
    ws_summary.cell(row=8, column=2).value = "Lista de Lojas com Divergência"
    ws_summary.cell(row=8, column=2).font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")

    tbl_headers = ["ID", "Loja", "Diferença R$"]
    for c_idx, h_text in enumerate(tbl_headers, 2):
        cell = ws_summary.cell(row=9, column=c_idx)
        cell.value = h_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    div_rows = df_merged[df_merged['Status'] == 'DIVERGENTE']
    tbl_row = 10
    if len(div_rows) == 0:
        ws_summary.cell(row=tbl_row, column=2).value = "Nenhuma divergência identificada! Ótimo dia de vendas."
        ws_summary.merge_cells(start_row=tbl_row, start_column=2, end_row=tbl_row, end_column=4)
        ws_summary.cell(row=tbl_row, column=2).font = Font(name="Segoe UI", size=10, italic=True)
        ws_summary.cell(row=tbl_row, column=2).alignment = Alignment(horizontal="center")
    else:
        for idx, row in div_rows.iterrows():
            ws_summary.cell(row=tbl_row, column=2).value = row['id_loja']
            ws_summary.cell(row=tbl_row, column=3).value = row['Loja_consinco']
            ws_summary.cell(row=tbl_row, column=4).value = row['Diferença']
            
            ws_summary.cell(row=tbl_row, column=2).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=tbl_row, column=3).alignment = Alignment(horizontal="left")
            ws_summary.cell(row=tbl_row, column=4).alignment = Alignment(horizontal="right")
            ws_summary.cell(row=tbl_row, column=4).number_format = 'R$ #,##0.00'
            
            for c in range(2, 5):
                cell = ws_summary.cell(row=tbl_row, column=c)
                cell.font = font_bold if c != 3 else font_body
                cell.border = thin_border
                cell.fill = fill_divergent_row
            tbl_row += 1

    # Larguras das colunas do painel
    ws_summary.column_dimensions['A'].width = 3
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 25
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 20
    ws_summary.column_dimensions['F'].width = 20

    # Salvando o resultado final
    saida_excel = "batimento_vendas_consolidado.xlsx"
    wb.save(saida_excel)
    print(f"\nSucesso! Arquivo gerado: '{saida_excel}'")

if __name__ == "__main__":
    executar_conciliacao()